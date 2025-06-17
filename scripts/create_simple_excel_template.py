#!/usr/bin/env python3
"""
Simple Excel Template Generator for Launch Transportation Management Platform
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_simple_excel_template():
    """Create a simple Excel workbook template."""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Create Instructions sheet first
    create_instructions_sheet(wb)
    
    # Create Drivers sheet
    create_drivers_sheet(wb)
    
    # Create Trucks sheet
    create_trucks_sheet(wb)
    
    # Create Trailers sheet
    create_trailers_sheet(wb)
    
    # Create Loads sheet
    create_loads_sheet(wb)
    
    # Create Settings sheet
    create_settings_sheet(wb)
    
    # Create Validation sheet
    create_validation_sheet(wb)
    
    # Save the workbook
    output_file = 'Launch_Import_Template.xlsx'
    wb.save(output_file)
    print(f"✅ Excel template created successfully: {output_file}")
    
    return output_file

def format_header_row(ws, row_num, color='1976D2'):
    """Format a header row with styling."""
    for cell in ws[row_num]:
        if cell.value:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

def create_drivers_sheet(wb):
    """Create the Drivers sheet."""
    ws = wb.create_sheet(title="Drivers")
    
    # Add description
    ws.merge_cells('A1:N1')
    ws['A1'] = "👨‍💼 Driver Information - Personal details, licensing, and emergency contacts"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    
    # Add headers
    headers = [
        'Driver ID*', 'First Name*', 'Last Name*', 'License Number*', 'License Expiry*',
        'Phone Number*', 'Email', 'Fuel Card Number', 'Assigned Truck ID', 'Status*',
        'Hire Date*', 'Emergency Contact Name*', 'Emergency Contact Phone*', 'Emergency Contact Relationship*'
    ]
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_header_row(ws, 3)
    
    # Add sample data
    sample_data = [
        ['DRV001', 'John', 'Smith', 'CDL123456789', '12/31/2025', '555-123-4567', 'john.smith@email.com', 'FC001', 'TRK001', 'active', '01/15/2023', 'Jane Smith', '555-987-6543', 'Spouse'],
        ['DRV002', 'Maria', 'Garcia', 'CDL987654321', '06/30/2026', '555-234-5678', 'maria.garcia@email.com', 'FC002', 'TRK002', 'active', '03/22/2023', 'Carlos Garcia', '555-876-5432', 'Spouse'],
        ['DRV003', 'Michael', 'Johnson', 'CDL456789123', '03/15/2025', '555-345-6789', 'mike.johnson@email.com', 'FC003', '', 'in-training', '11/10/2024', 'Sarah Johnson', '555-765-4321', 'Wife']
    ]
    
    for row in sample_data:
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

def create_trucks_sheet(wb):
    """Create the Trucks sheet."""
    ws = wb.create_sheet(title="Trucks")
    
    # Add description
    ws.merge_cells('A1:N1')
    ws['A1'] = "🚛 Truck Information - Vehicle specifications, maintenance records, and registration"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
    
    # Add headers
    headers = [
        'Truck ID*', 'Make*', 'Model*', 'Year*', 'License Plate*', 'VIN*', 'Color',
        'Assigned Driver ID', 'Status*', 'Mileage*', 'Last Maintenance Date',
        'Next Maintenance Due', 'Registration Expiry*', 'Insurance Expiry*'
    ]
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_header_row(ws, 3)
    
    # Add sample data
    sample_data = [
        ['TRK001', 'Freightliner', 'Cascadia', '2022', 'ABC123', '1FUJBBCK1NLAA1234', 'White', 'DRV001', 'in-use', '125000', '10/15/2024', '04/15/2025', '12/31/2025', '06/30/2025'],
        ['TRK002', 'Peterbilt', '579', '2021', 'DEF456', '1XP5DB9X1MD123456', 'Blue', 'DRV002', 'in-use', '98000', '11/20/2024', '05/20/2025', '11/30/2025', '08/15/2025'],
        ['TRK003', 'Kenworth', 'T680', '2023', 'GHI789', '1XKYD49X1PJ123456', 'Red', '', 'available', '45000', '12/01/2024', '06/01/2025', '01/31/2026', '09/30/2025']
    ]
    
    for row in sample_data:
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

def create_trailers_sheet(wb):
    """Create the Trailers sheet."""
    ws = wb.create_sheet(title="Trailers")
    
    # Add description
    ws.merge_cells('A1:O1')
    ws['A1'] = "🚚 Trailer Information - Specifications, capacity, and registration details"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    
    # Add headers
    headers = [
        'Trailer ID*', 'Make*', 'Model*', 'Year*', 'License Plate*', 'VIN*', 'Type*',
        'Capacity*', 'Length*', 'Assigned Truck ID', 'Status*', 'Last Maintenance Date',
        'Next Maintenance Due', 'Registration Expiry*', 'Insurance Expiry*'
    ]
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_header_row(ws, 3)
    
    # Add sample data
    sample_data = [
        ['TRL001', 'Great Dane', 'Super Seal', '2021', 'T123456', '1GRAA0625LB123456', 'Dry Van', '80000', '53', 'TRK001', 'in-use', '10/15/2024', '04/15/2025', '12/31/2025', '06/30/2025'],
        ['TRL002', 'Utility', '4000DX', '2022', 'T234567', '1UYVS2533MB123456', 'Dry Van', '80000', '53', 'TRK002', 'in-use', '11/20/2024', '05/20/2025', '11/30/2025', '08/15/2025'],
        ['TRL003', 'Wabash', 'Duraplate', '2020', 'T345678', '1JJV532W3KL123456', 'Dry Van', '80000', '53', '', 'available', '12/01/2024', '06/01/2025', '01/31/2026', '09/30/2025']
    ]
    
    for row in sample_data:
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

def create_loads_sheet(wb):
    """Create the Loads sheet."""
    ws = wb.create_sheet(title="Loads")
    
    # Add description
    ws.merge_cells('A1:V1')
    ws['A1'] = "📦 Load Information - Shipment details, locations, and assignments"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
    
    # Add headers
    headers = [
        'Load ID*', 'Load Number*', 'BOL Number', 'Shipper*', 'Pickup Address*', 'Pickup City*',
        'Pickup State*', 'Pickup Zip Code*', 'Delivery Address*', 'Delivery City*', 'Delivery State*',
        'Delivery Zip Code*', 'Assigned Driver ID', 'Assigned Truck ID', 'Assigned Trailer ID',
        'Status*', 'Cargo Description*', 'Weight*', 'Pickup Date*', 'Delivery Date*', 'Rate*', 'Notes'
    ]
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_header_row(ws, 3)
    
    # Add sample data
    sample_data = [
        ['LD001', 'L2024001', 'BOL123456', 'ABC Manufacturing', '123 Industrial Blvd', 'Dallas', 'TX', '75201', '456 Commerce St', 'Houston', 'TX', '77002', 'DRV001', 'TRK001', 'TRL001', 'in-transit', 'Auto Parts', '25000', '12/15/2024', '12/16/2024', '1250.00', 'Handle with care'],
        ['LD002', 'L2024002', 'BOL234567', 'XYZ Logistics', '789 Warehouse Dr', 'Atlanta', 'GA', '30309', '321 Distribution Way', 'Miami', 'FL', '33101', 'DRV002', 'TRK002', 'TRL002', 'assigned', 'Electronics', '18000', '12/16/2024', '12/18/2024', '1850.00', 'Temperature sensitive'],
        ['LD003', 'L2024003', 'BOL345678', 'Global Shipping', '555 Port Ave', 'Los Angeles', 'CA', '90012', '888 Freight Rd', 'Phoenix', 'AZ', '85001', '', '', '', 'pending', 'Textiles', '32000', '12/17/2024', '12/19/2024', '1650.00', '']
    ]
    
    for row in sample_data:
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

def create_settings_sheet(wb):
    """Create the Settings sheet."""
    ws = wb.create_sheet(title="Settings")
    
    # Add description
    ws.merge_cells('A1:D1')
    ws['A1'] = "⚙️ Company Settings - Configuration options and default values"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    
    # Add headers
    headers = ['Setting Name*', 'Setting Value*', 'Description', 'Category']
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_header_row(ws, 3)
    
    # Add sample data
    sample_data = [
        ['company_name', 'Launch Transportation Solutions', 'Official company name', 'Company'],
        ['company_address', '123 Logistics Way', 'Company headquarters address', 'Company'],
        ['company_phone', '555-LAUNCH', 'Main company phone number', 'Company'],
        ['default_currency', 'USD', 'Default currency for rates', 'System'],
        ['timezone', 'America/Chicago', 'Default timezone', 'System'],
        ['driver_id_prefix', 'DRV', 'Prefix for driver IDs', 'System'],
        ['truck_id_prefix', 'TRK', 'Prefix for truck IDs', 'System'],
        ['trailer_id_prefix', 'TRL', 'Prefix for trailer IDs', 'System'],
        ['load_id_prefix', 'LD', 'Prefix for load IDs', 'System']
    ]
    
    for row in sample_data:
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

def create_instructions_sheet(wb):
    """Create the Instructions sheet."""
    ws = wb.create_sheet(title="Instructions")
    
    instructions = [
        "🚀 Launch Transportation Management - Data Import Template",
        "",
        "📋 OVERVIEW",
        "This Excel workbook is designed for importing your existing transportation data",
        "into the Launch platform. Each sheet represents a different data type.",
        "",
        "📝 GENERAL GUIDELINES",
        "• Do NOT modify column headers - the system relies on exact column names",
        "• Required fields are marked with an asterisk (*) in the header",
        "• Date format should be MM/DD/YYYY",
        "• Phone numbers should include area code (e.g., 555-123-4567)",
        "• IDs should be unique within each sheet",
        "",
        "📊 SHEET DESCRIPTIONS",
        "",
        "👨‍💼 DRIVERS SHEET - Driver personal details, licensing, emergency contacts",
        "🚛 TRUCKS SHEET - Vehicle specifications, maintenance, registration",
        "🚚 TRAILERS SHEET - Trailer specs, capacity, registration details",
        "📦 LOADS SHEET - Shipment details, locations, assignments",
        "⚙️ SETTINGS SHEET - Company configuration and preferences",
        "",
        "🔍 VALID STATUS VALUES",
        "",
        "Driver Status: active, in-training, oos",
        "Truck/Trailer Status: available, in-use, maintenance, out-of-service",
        "Load Status: pending, assigned, picked-up, in-transit, delivered, cancelled",
        "",
        "📤 IMPORT PROCESS",
        "1. Complete all relevant sheets with your data",
        "2. Save as Excel format (.xlsx)",
        "3. Upload through Launch platform import feature",
        "4. Review import summary and resolve any errors",
        "5. Confirm import to add data to your system",
        "",
        f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ]
    
    # Add instructions to sheet
    for row_num, instruction in enumerate(instructions, 1):
        ws[f'A{row_num}'] = instruction
        
        # Format headers
        if instruction.startswith('🚀') or instruction.startswith('📋'):
            ws[f'A{row_num}'].font = Font(bold=True, size=14, color='1976D2')
        elif instruction.startswith('📝') or instruction.startswith('📊') or instruction.startswith('📤'):
            ws[f'A{row_num}'].font = Font(bold=True, size=12, color='1976D2')
        elif instruction.startswith('👨‍💼') or instruction.startswith('🚛') or instruction.startswith('🚚') or instruction.startswith('📦') or instruction.startswith('⚙️'):
            ws[f'A{row_num}'].font = Font(bold=True, size=11, color='D32F2F')
    
    # Set column width
    ws.column_dimensions['A'].width = 80

def create_validation_sheet(wb):
    """Create a validation sheet with reference data."""
    ws = wb.create_sheet(title="Validation")
    
    # Add title
    ws.merge_cells('A1:D1')
    ws['A1'] = "🔍 Data Validation Reference"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.append([])
    ws.append(['Status Type', 'Valid Values', 'Description', ''])
    format_header_row(ws, 3)
    
    # Add validation data
    validation_data = [
        ['Driver Status', 'active', 'Driver is actively working', ''],
        ['', 'in-training', 'Driver is in training period', ''],
        ['', 'oos', 'Driver is out of service', ''],
        ['', '', '', ''],
        ['Vehicle Status', 'available', 'Vehicle is available for assignment', ''],
        ['', 'in-use', 'Vehicle is currently assigned', ''],
        ['', 'maintenance', 'Vehicle is undergoing maintenance', ''],
        ['', 'out-of-service', 'Vehicle is out of service', ''],
        ['', '', '', ''],
        ['Load Status', 'pending', 'Load is awaiting assignment', ''],
        ['', 'assigned', 'Load has been assigned', ''],
        ['', 'picked-up', 'Load has been picked up', ''],
        ['', 'in-transit', 'Load is en route', ''],
        ['', 'delivered', 'Load has been delivered', ''],
        ['', 'cancelled', 'Load has been cancelled', '']
    ]
    
    for row in validation_data:
        ws.append(row)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35

if __name__ == "__main__":
    create_simple_excel_template()
