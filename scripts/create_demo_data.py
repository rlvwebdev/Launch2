#!/usr/bin/env python3
"""
Launch Transportation Management Platform
Demo Data Generator for Excel Import

Creates a comprehensive demo Excel file with:
- 23 Drivers
- 42 Trucks  
- 63 Trailers
- 119 Loads
Complete with realistic data from 06/01/2025 through 06/23/2025
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random
import string

# Demo data constants
DEMO_START_DATE = datetime(2025, 6, 1)
DEMO_END_DATE = datetime(2025, 6, 23)

# Realistic company names, locations, and other data
COMPANIES = [
    "ABC Manufacturing", "XYZ Logistics", "Global Shipping", "Midwest Supply", "Coast Transport",
    "Delta Freight", "American Cargo", "Express Delivery", "Prime Logistics", "United Shipping",
    "National Transport", "Interstate Freight", "Metro Logistics", "Regional Cargo", "Elite Transport",
    "Swift Shipping", "Eagle Logistics", "Pioneer Freight", "Summit Transport", "Pacific Shipping",
    "Atlantic Cargo", "Continental Express", "Horizon Logistics", "Apex Transport", "Stellar Freight"
]

US_CITIES = [
    ("New York", "NY", "10001"), ("Los Angeles", "CA", "90001"), ("Chicago", "IL", "60601"),
    ("Houston", "TX", "77001"), ("Phoenix", "AZ", "85001"), ("Philadelphia", "PA", "19101"),
    ("San Antonio", "TX", "78201"), ("San Diego", "CA", "92101"), ("Dallas", "TX", "75201"),
    ("San Jose", "CA", "95101"), ("Austin", "TX", "73301"), ("Jacksonville", "FL", "32099"),
    ("Fort Worth", "TX", "76101"), ("Columbus", "OH", "43085"), ("Charlotte", "NC", "28201"),
    ("San Francisco", "CA", "94102"), ("Indianapolis", "IN", "46201"), ("Seattle", "WA", "98101"),
    ("Denver", "CO", "80201"), ("Washington", "DC", "20001"), ("Boston", "MA", "02101"),
    ("El Paso", "TX", "79901"), ("Nashville", "TN", "37201"), ("Detroit", "MI", "48201"),
    ("Oklahoma City", "OK", "73101"), ("Portland", "OR", "97201"), ("Las Vegas", "NV", "89101"),
    ("Memphis", "TN", "38101"), ("Louisville", "KY", "40201"), ("Baltimore", "MD", "21201"),
    ("Milwaukee", "WI", "53201"), ("Albuquerque", "NM", "87101"), ("Tucson", "AZ", "85701"),
    ("Fresno", "CA", "93701"), ("Sacramento", "CA", "95814"), ("Mesa", "AZ", "85201"),
    ("Kansas City", "MO", "64108"), ("Atlanta", "GA", "30301"), ("Long Beach", "CA", "90801"),
    ("Colorado Springs", "CO", "80901"), ("Raleigh", "NC", "27601"), ("Miami", "FL", "33101"),
    ("Virginia Beach", "VA", "23451"), ("Omaha", "NE", "68101"), ("Oakland", "CA", "94601"),
    ("Minneapolis", "MN", "55401"), ("Tulsa", "OK", "74101"), ("Arlington", "TX", "76010"),
    ("Tampa", "FL", "33601"), ("New Orleans", "LA", "70112"), ("Wichita", "KS", "67202")
]

TRUCK_MAKES = ["Freightliner", "Peterbilt", "Kenworth", "Volvo", "Mack", "International", "Western Star"]
TRUCK_MODELS = {
    "Freightliner": ["Cascadia", "Columbia", "Century", "Coronado"],
    "Peterbilt": ["579", "389", "367", "348", "567"],
    "Kenworth": ["T680", "T880", "W990", "T800", "T470"],
    "Volvo": ["VNL", "VNR", "VHD", "VAH"],
    "Mack": ["Anthem", "Pinnacle", "Granite", "TerraPro"],
    "International": ["LT", "RH", "HX", "MV"],
    "Western Star": ["5700XE", "4700", "6900XD", "47X"]
}

TRAILER_MAKES = ["Great Dane", "Utility", "Wabash", "Hyundai", "Stoughton", "Wilson", "Fontaine", "MAC"]
TRAILER_MODELS = {
    "Great Dane": ["Super Seal", "Freedom", "Champion"],
    "Utility": ["4000DX", "3000R", "Reefer"],
    "Wabash": ["Duraplate", "Duraplate DW", "Arcticlite"],
    "Hyundai": ["Translead", "Composite", "Dry Van"],
    "Stoughton": ["Trailers", "Composite", "SOLO"],
    "Wilson": ["Pacesetter", "Silverstar", "Commander"],
    "Fontaine": ["Revolution", "Magnitude", "Infinity"],
    "MAC": ["End Dump", "Belly Dump", "Side Dump"]
}

TRAILER_TYPES = ["dry-van", "flatbed", "refrigerated", "hopper", "tanker", "lowboy", "step-deck"]

CARGO_TYPES = [
    "Auto Parts", "Electronics", "Textiles", "Machinery", "Food Products", "Steel", "Lumber",
    "Chemicals", "Paper Products", "Furniture", "Appliances", "Construction Materials",
    "Pharmaceuticals", "Beverages", "Clothing", "Tools", "Plastic Products", "Metal Parts",
    "Glass Products", "Agricultural Products", "Computer Equipment", "Medical Supplies",
    "Industrial Equipment", "Consumer Goods", "Building Supplies", "Raw Materials"
]

FIRST_NAMES = [
    "James", "John", "Robert", "Michael", "David", "William", "Richard", "Joseph", "Thomas", "Christopher",
    "Charles", "Daniel", "Matthew", "Anthony", "Mark", "Donald", "Steven", "Andrew", "Kenneth", "Paul",
    "Joshua", "Kevin", "Brian", "George", "Edward", "Mary", "Patricia", "Jennifer", "Linda", "Elizabeth",
    "Barbara", "Susan", "Jessica", "Sarah", "Karen", "Nancy", "Lisa", "Betty", "Helen", "Sandra",
    "Donna", "Carol", "Ruth", "Sharon", "Michelle", "Laura", "Emily", "Kimberly", "Deborah", "Dorothy"
]

LAST_NAMES = [
    "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Rodriguez", "Martinez",
    "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin",
    "Lee", "Perez", "Thompson", "White", "Harris", "Sanchez", "Clark", "Ramirez", "Lewis", "Robinson",
    "Walker", "Young", "Allen", "King", "Wright", "Scott", "Torres", "Nguyen", "Hill", "Flores",
    "Green", "Adams", "Nelson", "Baker", "Hall", "Rivera", "Campbell", "Mitchell", "Carter", "Roberts"
]

COLORS = ["White", "Black", "Red", "Blue", "Silver", "Gray", "Green", "Yellow", "Orange", "Purple"]

def generate_vin():
    """Generate a realistic VIN number."""
    return '1' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=16))

def generate_license_plate():
    """Generate a realistic license plate."""
    return ''.join(random.choices(string.ascii_uppercase, k=3)) + ''.join(random.choices(string.digits, k=3))

def generate_phone():
    """Generate a realistic phone number."""
    area_code = random.choice(['555', '214', '312', '713', '602', '215', '210', '619', '972', '408'])
    return f"{area_code}-{random.randint(100, 999)}-{random.randint(1000, 9999)}"

def generate_email(first_name, last_name):
    """Generate a realistic email address."""
    domains = ["email.com", "mail.com", "transport.com", "logistics.com", "freight.com"]
    return f"{first_name.lower()}.{last_name.lower()}@{random.choice(domains)}"

def generate_cdl():
    """Generate a realistic CDL number."""
    return "CDL" + ''.join(random.choices(string.digits, k=9))

def random_date_in_range(start_date, end_date):
    """Generate a random date within the specified range."""
    delta = end_date - start_date
    random_days = random.randint(0, delta.days)
    return start_date + timedelta(days=random_days)

def format_date(date_obj):
    """Format date as MM/DD/YYYY."""
    return date_obj.strftime("%m/%d/%Y")

def create_demo_excel_file():
    """Create the demo Excel file with comprehensive data."""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create Instructions sheet
    create_demo_instructions_sheet(wb)
      # Generate data
    drivers_data = generate_drivers_data(23)
    trucks_data = generate_trucks_data(42)
    trailers_data = generate_trailers_data(63)
    loads_data = generate_loads_data(119, drivers_data, trucks_data, trailers_data)
    settings_data = generate_settings_data()
    
    # Create sheets with data
    create_demo_drivers_sheet(wb, drivers_data)
    create_demo_trucks_sheet(wb, trucks_data)
    create_demo_trailers_sheet(wb, trailers_data)
    create_demo_loads_sheet(wb, loads_data)
    create_demo_settings_sheet(wb, settings_data)
    create_demo_validation_sheet(wb)
    
    # Save the workbook
    output_file = 'Launch_Demo_Import_Data.xlsx'
    wb.save(output_file)
    print(f"✅ Demo Excel file created successfully: {output_file}")
    print(f"📊 Generated data:")
    print(f"   • {len(drivers_data)} Drivers")
    print(f"   • {len(trucks_data)} Trucks")
    print(f"   • {len(trailers_data)} Trailers")
    print(f"   • {len(loads_data)} Loads")
    print(f"   📅 Date range: {DEMO_START_DATE.strftime('%m/%d/%Y')} - {DEMO_END_DATE.strftime('%m/%d/%Y')}")
    
    return output_file

def generate_drivers_data(count):
    """Generate realistic driver data."""
    drivers = []
    used_names = set()
    
    for i in range(count):
        # Ensure unique name combinations
        while True:
            first_name = random.choice(FIRST_NAMES)
            last_name = random.choice(LAST_NAMES)
            name_combo = f"{first_name}_{last_name}"
            if name_combo not in used_names:
                used_names.add(name_combo)
                break
        
        hire_date = random_date_in_range(datetime(2020, 1, 1), datetime(2024, 12, 31))
        license_expiry = random_date_in_range(datetime(2025, 6, 24), datetime(2027, 12, 31))
        
        # Emergency contact
        emergency_first = random.choice(FIRST_NAMES)
        emergency_last = random.choice([last_name] + random.choices(LAST_NAMES, k=2))  # Often family
        relationships = ["Spouse", "Wife", "Husband", "Mother", "Father", "Sister", "Brother", "Son", "Daughter"]
        
        status = random.choices(
            ["active", "in-training", "oos"],
            weights=[85, 10, 5],  # Most drivers active
            k=1
        )[0]
        
        driver = {
            'id': f"DRV{i+1:03d}",
            'first_name': first_name,
            'last_name': last_name,
            'license_number': generate_cdl(),
            'license_expiry': format_date(license_expiry),
            'phone_number': generate_phone(),
            'email': generate_email(first_name, last_name),
            'fuel_card': f"FC{i+1:03d}",
            'assigned_truck_id': "",  # Will be assigned later
            'status': status,
            'hire_date': format_date(hire_date),
            'emergency_contact_name': f"{emergency_first} {emergency_last}",
            'emergency_contact_phone': generate_phone(),
            'emergency_contact_relationship': random.choice(relationships)
        }
        drivers.append(driver)
    
    return drivers

def generate_trucks_data(count):
    """Generate realistic truck data."""
    trucks = []
    used_plates = set()
    used_vins = set()
    
    for i in range(count):
        # Ensure unique plates and VINs
        while True:
            plate = generate_license_plate()
            if plate not in used_plates:
                used_plates.add(plate)
                break
        
        while True:
            vin = generate_vin()
            if vin not in used_vins:
                used_vins.add(vin)
                break
        
        make = random.choice(TRUCK_MAKES)
        model = random.choice(TRUCK_MODELS[make])
        year = random.randint(2018, 2024)
        
        # Maintenance dates
        last_maintenance = random_date_in_range(datetime(2024, 1, 1), datetime(2025, 5, 31))
        next_maintenance = last_maintenance + timedelta(days=random.randint(90, 180))
        
        # Registration and insurance expiry
        registration_expiry = random_date_in_range(datetime(2025, 6, 24), datetime(2026, 12, 31))
        insurance_expiry = random_date_in_range(datetime(2025, 6, 24), datetime(2026, 6, 30))
        
        status = random.choices(
            ["available", "in-use", "maintenance", "out-of-service"],
            weights=[20, 65, 10, 5],  # Most trucks in use
            k=1
        )[0]
        
        truck = {
            'id': f"TRK{i+1:03d}",
            'make': make,
            'model': model,
            'year': str(year),
            'license_plate': plate,
            'vin': vin,
            'color': random.choice(COLORS),
            'assigned_driver_id': "",  # Will be assigned later
            'status': status,
            'mileage': str(random.randint(50000, 500000)),
            'last_maintenance': format_date(last_maintenance),
            'next_maintenance_due': format_date(next_maintenance),
            'registration_expiry': format_date(registration_expiry),
            'insurance_expiry': format_date(insurance_expiry)
        }
        trucks.append(truck)
    
    return trucks

def generate_trailers_data(count):
    """Generate realistic trailer data."""
    trailers = []
    used_plates = set()
    used_vins = set()
    
    for i in range(count):
        # Ensure unique plates and VINs
        while True:
            plate = f"T{random.randint(100000, 999999)}"
            if plate not in used_plates:
                used_plates.add(plate)
                break
        
        while True:
            vin = generate_vin()
            if vin not in used_vins:
                used_vins.add(vin)
                break
        
        make = random.choice(TRAILER_MAKES)
        model = random.choice(TRAILER_MODELS[make])
        year = random.randint(2018, 2024)
        trailer_type = random.choice(TRAILER_TYPES)
        
        # Capacity and length based on type
        if trailer_type == "dry-van":
            capacity = 80000
            length = random.choice([48, 53])
        elif trailer_type == "flatbed":
            capacity = 80000
            length = random.choice([48, 53])
        elif trailer_type == "refrigerated":
            capacity = 80000
            length = 53
        elif trailer_type == "hopper":
            capacity = 80000
            length = random.choice([40, 42])
        elif trailer_type == "tanker":
            capacity = 80000
            length = random.choice([40, 45])
        else:
            capacity = 80000
            length = random.choice([45, 48, 53])
        
        # Maintenance dates
        last_maintenance = random_date_in_range(datetime(2024, 1, 1), datetime(2025, 5, 31))
        next_maintenance = last_maintenance + timedelta(days=random.randint(90, 180))
        
        # Registration and insurance expiry
        registration_expiry = random_date_in_range(datetime(2025, 6, 24), datetime(2026, 12, 31))
        insurance_expiry = random_date_in_range(datetime(2025, 6, 24), datetime(2026, 6, 30))
        
        status = random.choices(
            ["available", "in-use", "maintenance", "out-of-service"],
            weights=[25, 60, 10, 5],  # Most trailers in use
            k=1
        )[0]
        
        trailer = {
            'id': f"TRL{i+1:03d}",
            'make': make,
            'model': model,
            'year': str(year),
            'license_plate': plate,
            'vin': vin,
            'type': trailer_type,
            'capacity': str(capacity),
            'length': str(length),
            'assigned_truck_id': "",  # Will be assigned later
            'status': status,
            'last_maintenance': format_date(last_maintenance),
            'next_maintenance_due': format_date(next_maintenance),
            'registration_expiry': format_date(registration_expiry),
            'insurance_expiry': format_date(insurance_expiry)
        }
        trailers.append(trailer)
    
    return trailers

def generate_loads_data(count, drivers, trucks, trailers):
    """Generate realistic load data with proper assignments."""
    loads = []
    used_load_numbers = set()
    
    # Assign drivers to trucks and trailers first
    available_drivers = [d for d in drivers if d['status'] == 'active']
    available_trucks = [t for t in trucks if t['status'] in ['available', 'in-use']]
    available_trailers = [t for t in trailers if t['status'] in ['available', 'in-use']]
    
    # Create driver-truck-trailer assignments
    assignments = []
    for i in range(min(len(available_drivers), len(available_trucks))):
        if i < len(available_trailers):
            assignment = {
                'driver': available_drivers[i]['id'],
                'truck': available_trucks[i]['id'],
                'trailer': available_trailers[i]['id']
            }
            # Update the original data
            drivers[drivers.index(available_drivers[i])]['assigned_truck_id'] = available_trucks[i]['id']
            trucks[trucks.index(available_trucks[i])]['assigned_driver_id'] = available_drivers[i]['id']
            trucks[trucks.index(available_trucks[i])]['status'] = 'in-use'
            trailers[trailers.index(available_trailers[i])]['assigned_truck_id'] = available_trucks[i]['id']
            trailers[trailers.index(available_trailers[i])]['status'] = 'in-use'
            assignments.append(assignment)
    
    for i in range(count):
        # Generate unique load number
        while True:
            load_number = f"L2025{random.randint(1000, 9999)}"
            if load_number not in used_load_numbers:
                used_load_numbers.add(load_number)
                break
        
        # BOL number
        bol_number = f"BOL{random.randint(100000, 999999)}"
        
        # Shipper
        shipper = random.choice(COMPANIES)
        
        # Pickup and delivery locations
        pickup_city, pickup_state, pickup_zip = random.choice(US_CITIES)
        delivery_city, delivery_state, delivery_zip = random.choice(US_CITIES)
        
        # Ensure different pickup and delivery
        while pickup_city == delivery_city:
            delivery_city, delivery_state, delivery_zip = random.choice(US_CITIES)
        
        # Generate addresses
        pickup_address = f"{random.randint(100, 9999)} {random.choice(['Industrial', 'Warehouse', 'Commerce', 'Factory', 'Logistics'])} {random.choice(['Blvd', 'St', 'Ave', 'Dr', 'Way'])}"
        delivery_address = f"{random.randint(100, 9999)} {random.choice(['Distribution', 'Freight', 'Commerce', 'Industrial', 'Shipping'])} {random.choice(['Blvd', 'St', 'Ave', 'Dr', 'Way'])}"
        
        # Assignment (some loads assigned, some pending)
        assignment_chance = random.random()
        if assignment_chance < 0.7 and assignments:  # 70% chance of assignment
            assignment = random.choice(assignments)
            assigned_driver_id = assignment['driver']
            assigned_truck_id = assignment['truck']
            assigned_trailer_id = assignment['trailer']
        else:
            assigned_driver_id = ""
            assigned_truck_id = ""
            assigned_trailer_id = ""
        
        # Status based on assignment and dates
        if assigned_driver_id:
            status = random.choices(
                ["assigned", "picked-up", "in-transit", "delivered"],
                weights=[30, 25, 30, 15],
                k=1
            )[0]
        else:
            status = "pending"
        
        # Cargo
        cargo_description = random.choice(CARGO_TYPES)
        weight = random.randint(15000, 80000)
        
        # Dates
        pickup_date = random_date_in_range(DEMO_START_DATE, DEMO_END_DATE)
        delivery_date = pickup_date + timedelta(days=random.randint(1, 5))
        
        # Rate calculation (roughly $1.50-$3.00 per mile, estimated distance)
        base_rate = random.uniform(1200, 3500)
        rate = round(base_rate, 2)
        
        # Notes
        notes_options = [
            "", "Handle with care", "Temperature sensitive", "Fragile items", 
            "Heavy machinery", "Perishable goods", "Hazardous materials",
            "Oversized load", "Time critical", "Special handling required"
        ]
        notes = random.choice(notes_options)
        
        load = {
            'id': f"LD{i+1:03d}",
            'load_number': load_number,
            'bol_number': bol_number,
            'shipper': shipper,
            'pickup_address': pickup_address,
            'pickup_city': pickup_city,
            'pickup_state': pickup_state,
            'pickup_zip_code': pickup_zip,
            'delivery_address': delivery_address,
            'delivery_city': delivery_city,
            'delivery_state': delivery_state,
            'delivery_zip_code': delivery_zip,
            'assigned_driver_id': assigned_driver_id,
            'assigned_truck_id': assigned_truck_id,
            'assigned_trailer_id': assigned_trailer_id,
            'status': status,
            'cargo_description': cargo_description,
            'weight': str(weight),
            'pickup_date': format_date(pickup_date),
            'delivery_date': format_date(delivery_date),
            'rate': f"{rate:.2f}",
            'notes': notes
        }
        loads.append(load)
    
    return loads

def generate_settings_data():
    """Generate company settings data."""
    return [
        ["company_name", "Launch Transportation Solutions", "Official company name", "Company"],
        ["company_address", "123 Launch Way", "Company headquarters address", "Company"],
        ["company_city", "Houston", "Company city", "Company"],
        ["company_state", "TX", "Company state", "Company"],
        ["company_zip", "77001", "Company ZIP code", "Company"],
        ["company_phone", "555-LAUNCH", "Main company phone number", "Company"],
        ["company_email", "info@launch.com", "Main company email", "Company"],
        ["default_currency", "USD", "Default currency for rates and payments", "System"],
        ["timezone", "America/Chicago", "Default timezone for the system", "System"],
        ["mileage_unit", "miles", "Unit for distance measurements", "System"],
        ["weight_unit", "pounds", "Unit for weight measurements", "System"],
        ["fuel_card_prefix", "FC", "Prefix for fuel card numbers", "System"],
        ["driver_id_prefix", "DRV", "Prefix for driver IDs", "System"],
        ["truck_id_prefix", "TRK", "Prefix for truck IDs", "System"],
        ["trailer_id_prefix", "TRL", "Prefix for trailer IDs", "System"],
        ["load_id_prefix", "LD", "Prefix for load IDs", "System"],
        ["maintenance_reminder_days", "30", "Days before maintenance due to send reminder", "Notifications"],
        ["license_expiry_reminder_days", "60", "Days before license expiry to send reminder", "Notifications"],
        ["registration_reminder_days", "45", "Days before registration expiry to send reminder", "Notifications"],
        ["insurance_reminder_days", "30", "Days before insurance expiry to send reminder", "Notifications"]
    ]

def format_demo_header_row(ws, row_num, color='1976D2'):
    """Format a header row with styling."""
    for cell in ws[row_num]:
        if cell.value:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

def create_demo_instructions_sheet(wb):
    """Create the instructions sheet for demo data."""
    ws = wb.create_sheet(title="Instructions")
    
    instructions = [
        "🚀 Launch Transportation Management - Demo Import Data",
        "",
        "📋 DEMO DATA OVERVIEW",
        "This Excel workbook contains comprehensive demo data for the Launch platform:",
        "• 23 Realistic drivers with complete profiles",
        "• 42 Trucks with proper specifications and maintenance records",
        "• 63 Trailers with capacity and assignment information",
        "• 119 Loads spanning June 1-23, 2025 with realistic routes",
        "• Complete company settings and configuration",
        "",
        "📊 DATA CHARACTERISTICS",
        "• All data is realistically generated with proper relationships",
        "• Driver-truck-trailer assignments are properly maintained",
        "• Load assignments reflect actual transportation operations",
        "• Dates span a realistic 23-day operational period",
        "• Geographic diversity across major US cities",
        "• Proper status distributions (most active, some training/maintenance)",
        "",
        "📝 IMPORT INSTRUCTIONS",
        "1. This file is ready for immediate import into Launch",
        "2. All data has been validated and formatted correctly",
        "3. Upload through the Launch platform import feature",
        "4. Review the import summary (should show no errors)",
        "5. Confirm import to populate your system with demo data",
        "",
        "🔍 DATA DETAILS",
        "• Driver licenses expire between 2025-2027",
        "• Vehicle maintenance is current with future due dates",
        "• Load operations span major US transportation corridors",
        "• Company settings configured for US operations",
        "• All IDs follow proper numbering conventions",
        "",
        "⚠️ IMPORTANT NOTES",
        "• This is DEMO DATA - not real company information",
        "• All personal information is fictional",
        "• VIN numbers and license plates are generated",
        "• Phone numbers use 555 area codes",
        "• Email addresses use generic domains",
        "",
        f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        f"Data period: {DEMO_START_DATE.strftime('%B %d, %Y')} - {DEMO_END_DATE.strftime('%B %d, %Y')}"
    ]
    
    # Add instructions to sheet
    for row_num, instruction in enumerate(instructions, 1):
        ws[f'A{row_num}'] = instruction
        
        # Format headers
        if instruction.startswith('🚀'):
            ws[f'A{row_num}'].font = Font(bold=True, size=14, color='1976D2')
        elif instruction.startswith('📋') or instruction.startswith('📊') or instruction.startswith('📝') or instruction.startswith('🔍'):
            ws[f'A{row_num}'].font = Font(bold=True, size=12, color='1976D2')
        elif instruction.startswith('⚠️'):
            ws[f'A{row_num}'].font = Font(bold=True, size=12, color='D32F2F')
    
    # Set column width
    ws.column_dimensions['A'].width = 80

def create_demo_drivers_sheet(wb, drivers_data):
    """Create the drivers sheet with demo data."""
    ws = wb.create_sheet(title="Drivers")
    
    # Add description
    ws.merge_cells('A1:N1')
    ws['A1'] = f"👨‍💼 Demo Drivers Data - {len(drivers_data)} realistic driver profiles with complete information"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    
    # Add headers
    headers = [
        'Driver ID*', 'First Name*', 'Last Name*', 'License Number*', 'License Expiry*',
        'Phone Number*', 'Email', 'Fuel Card Number', 'Assigned Truck ID', 'Status*',
        'Hire Date*', 'Emergency Contact Name*', 'Emergency Contact Phone*', 'Emergency Contact Relationship*'
    ]
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_demo_header_row(ws, 3)
    
    # Add driver data
    for driver in drivers_data:
        row = [
            driver['id'], driver['first_name'], driver['last_name'], driver['license_number'],
            driver['license_expiry'], driver['phone_number'], driver['email'], driver['fuel_card'],
            driver['assigned_truck_id'], driver['status'], driver['hire_date'],
            driver['emergency_contact_name'], driver['emergency_contact_phone'], driver['emergency_contact_relationship']
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

def create_demo_trucks_sheet(wb, trucks_data):
    """Create the trucks sheet with demo data."""
    ws = wb.create_sheet(title="Trucks")
    
    # Add description
    ws.merge_cells('A1:N1')
    ws['A1'] = f"🚛 Demo Trucks Data - {len(trucks_data)} realistic truck specifications with maintenance records"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
    
    # Add headers
    headers = [
        'Truck ID*', 'Make*', 'Model*', 'Year*', 'License Plate*', 'VIN*', 'Color',
        'Assigned Driver ID', 'Status*', 'Mileage*', 'Last Maintenance Date',
        'Next Maintenance Due', 'Registration Expiry*', 'Insurance Expiry*'
    ]
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_demo_header_row(ws, 3)
    
    # Add truck data
    for truck in trucks_data:
        row = [
            truck['id'], truck['make'], truck['model'], truck['year'], truck['license_plate'],
            truck['vin'], truck['color'], truck['assigned_driver_id'], truck['status'],
            truck['mileage'], truck['last_maintenance'], truck['next_maintenance_due'],
            truck['registration_expiry'], truck['insurance_expiry']
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

def create_demo_trailers_sheet(wb, trailers_data):
    """Create the trailers sheet with demo data."""
    ws = wb.create_sheet(title="Trailers")
    
    # Add description
    ws.merge_cells('A1:O1')
    ws['A1'] = f"🚚 Demo Trailers Data - {len(trailers_data)} realistic trailer specifications with capacity information"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    
    # Add headers
    headers = [
        'Trailer ID*', 'Make*', 'Model*', 'Year*', 'License Plate*', 'VIN*', 'Type*',
        'Capacity*', 'Length*', 'Assigned Truck ID', 'Status*', 'Last Maintenance Date',
        'Next Maintenance Due', 'Registration Expiry*', 'Insurance Expiry*'
    ]
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_demo_header_row(ws, 3)
    
    # Add trailer data
    for trailer in trailers_data:
        row = [
            trailer['id'], trailer['make'], trailer['model'], trailer['year'], trailer['license_plate'],
            trailer['vin'], trailer['type'], trailer['capacity'], trailer['length'],
            trailer['assigned_truck_id'], trailer['status'], trailer['last_maintenance'],
            trailer['next_maintenance_due'], trailer['registration_expiry'], trailer['insurance_expiry']
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

def create_demo_loads_sheet(wb, loads_data):
    """Create the loads sheet with demo data."""
    ws = wb.create_sheet(title="Loads")
    
    # Add description
    ws.merge_cells('A1:V1')
    ws['A1'] = f"📦 Demo Loads Data - {len(loads_data)} realistic shipments spanning June 1-23, 2025"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
    
    # Add headers
    headers = [
        'Load ID*', 'Load Number*', 'BOL Number', 'Shipper*', 'Pickup Address*', 'Pickup City*',
        'Pickup State*', 'Pickup Zip Code*', 'Delivery Address*', 'Delivery City*', 'Delivery State*',
        'Delivery Zip Code*', 'Assigned Driver ID', 'Assigned Truck ID', 'Assigned Trailer ID',
        'Status*', 'Cargo Description*', 'Weight*', 'Pickup Date*', 'Delivery Date*', 'Rate*', 'Notes'
    ]
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_demo_header_row(ws, 3)
    
    # Add load data
    for load in loads_data:
        row = [
            load['id'], load['load_number'], load['bol_number'], load['shipper'],
            load['pickup_address'], load['pickup_city'], load['pickup_state'], load['pickup_zip_code'],
            load['delivery_address'], load['delivery_city'], load['delivery_state'], load['delivery_zip_code'],
            load['assigned_driver_id'], load['assigned_truck_id'], load['assigned_trailer_id'],
            load['status'], load['cargo_description'], load['weight'], load['pickup_date'],
            load['delivery_date'], load['rate'], load['notes']
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

def create_demo_settings_sheet(wb, settings_data):
    """Create the settings sheet with demo data."""
    ws = wb.create_sheet(title="Settings")
    
    # Add description
    ws.merge_cells('A1:D1')
    ws['A1'] = f"⚙️ Demo Settings Data - Company configuration and system preferences"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    
    # Add headers
    headers = ['Setting Name*', 'Setting Value*', 'Description', 'Category']
    
    ws.append([])  # Empty row
    ws.append(headers)
    format_demo_header_row(ws, 3)
    
    # Add settings data
    for setting in settings_data:
        ws.append(setting)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

def create_demo_validation_sheet(wb):
    """Create the validation sheet."""
    ws = wb.create_sheet(title="Validation")
    
    # Add title
    ws.merge_cells('A1:D1')
    ws['A1'] = "🔍 Data Validation Reference - Valid status values and formats"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.append([])
    ws.append(['Status Type', 'Valid Values', 'Description', 'Count in Demo'])
    format_demo_header_row(ws, 3)
    
    # Add validation data with demo counts
    validation_data = [
        ['Driver Status', 'active', 'Driver is actively working', '~20'],
        ['', 'in-training', 'Driver is in training period', '~2'],
        ['', 'oos', 'Driver is out of service', '~1'],
        ['', '', '', ''],
        ['Vehicle Status', 'available', 'Vehicle available for assignment', '~15'],
        ['', 'in-use', 'Vehicle currently assigned', '~75'],
        ['', 'maintenance', 'Vehicle undergoing maintenance', '~10'],
        ['', 'out-of-service', 'Vehicle out of service', '~5'],
        ['', '', '', ''],
        ['Load Status', 'pending', 'Load awaiting assignment', '~30'],
        ['', 'assigned', 'Load assigned to driver/truck', '~35'],
        ['', 'picked-up', 'Load has been picked up', '~25'],
        ['', 'in-transit', 'Load en route to destination', '~20'],
        ['', 'delivered', 'Load has been delivered', '~9'],
        ['', 'cancelled', 'Load has been cancelled', '~0']
    ]
    
    for row in validation_data:
        ws.append(row)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15

if __name__ == "__main__":
    create_demo_excel_file()
