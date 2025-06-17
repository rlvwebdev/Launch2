#!/usr/bin/env python3
"""
Launch Transportation Management Platform
Excel Template Generator

This script creates a comprehensive Excel workbook template for importing data
into the Launch transportation management system.
"""

import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_template():
    """Create the Launch import template Excel workbook."""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Define template directory
    template_dir = './templates'
    
    # Define sheet configurations
    sheets_config = {
        'Drivers': {
            'file': 'drivers_template.csv',
            'color': 'E3F2FD',  # Light Blue
            'description': 'Driver information including personal details, licensing, and emergency contacts'
        },
        'Trucks': {
            'file': 'trucks_template.csv',
            'color': 'E8F5E8',  # Light Green
            'description': 'Truck/vehicle information including specifications and maintenance records'
        },
        'Trailers': {
            'file': 'trailers_template.csv',
            'color': 'FFF3E0',  # Light Orange
            'description': 'Trailer information including specifications and registration details'
        },
        'Loads': {
            'file': 'loads_template.csv',
            'color': 'F3E5F5',  # Light Purple
            'description': 'Load/shipment information including pickup/delivery details and assignments'
        },
        'Settings': {
            'file': 'settings_template.csv',
            'color': 'FFF8E1',  # Light Yellow
            'description': 'Company settings, default values, and configuration options'
        }
    }
    
    # Create sheets from CSV data
    for sheet_name, config in sheets_config.items():
        csv_path = os.path.join(template_dir, config['file'])
        
        if os.path.exists(csv_path):
            # Read CSV data
            df = pd.read_csv(csv_path)
            
            # Create new sheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Add description row at the top
            ws.merge_cells('A1:Z1')
            ws['A1'] = f"📋 {config['description']}"
            ws['A1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A1'].fill = PatternFill(start_color=config['color'], end_color=config['color'], fill_type='solid')
            
            # Add empty row
            ws.append([])
            
            # Add data starting from row 3
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format header row (row 3)
            header_row = 3
            for col_num, cell in enumerate(ws[header_row], 1):
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
              # Auto-adjust column widths
            for col_num in range(1, ws.max_column + 1):
                column_letter = openpyxl.utils.get_column_letter(col_num)
                max_length = 0
                
                for row_num in range(3, ws.max_row + 1):  # Start from row 3 (data starts there)
                    cell = ws[f'{column_letter}{row_num}']
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = max(adjusted_width, 12)  # Minimum width of 12
            
            # Add borders to data rows
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
    
    # Create Instructions sheet
    create_instructions_sheet(wb)
    
    # Create Validation sheet
    create_validation_sheet(wb)
    
    # Move Instructions to be the first sheet
    wb.move_sheet(wb['Instructions'], offset=-len(wb.sheetnames)+1)
    
    # Save the workbook
    output_file = 'Launch_Import_Template.xlsx'
    wb.save(output_file)
    print(f"✅ Excel template created successfully: {output_file}")
    
    return output_file

def create_instructions_sheet(wb):
    """Create the instructions sheet with detailed guidance."""
    ws = wb.create_sheet(title="Instructions")
    
    # Instructions content
    instructions = [
        "🚀 Launch Transportation Management - Data Import Template",
        "",
        "📋 OVERVIEW",
        "This Excel workbook is designed for importing your existing transportation data",
        "into the Launch platform. Each sheet represents a different data type.",
        "",
        "📝 GENERAL GUIDELINES",
        "• Do NOT modify column headers - the system relies on exact column names",
        "• Required fields are marked with an asterisk (*) in the header",
        "• Date format should be MM/DD/YYYY",
        "• Phone numbers should include area code (e.g., 555-123-4567)",
        "• IDs should be unique within each sheet",
        "• Status values must match exactly (case-sensitive)",
        "",
        "📊 SHEET DESCRIPTIONS",
        "",
        "🧑‍💼 DRIVERS SHEET",
        "Contains driver information including:",
        "• Personal details (name, contact info)",
        "• License information and expiry dates",
        "• Emergency contact details",
        "• Current truck assignments",
        "• Employment status and hire date",
        "",
        "🚛 TRUCKS SHEET",
        "Contains truck/vehicle information including:",
        "• Vehicle specifications (make, model, year)",
        "• Registration and VIN details",
        "• Maintenance records and schedules",
        "• Current driver assignments",
        "• Insurance and registration expiry dates",
        "",
        "🚚 TRAILERS SHEET",
        "Contains trailer information including:",
        "• Trailer specifications and capacity",
        "• Registration and identification details",
        "• Maintenance schedules",
        "• Current truck assignments",
        "",
        "📦 LOADS SHEET",
        "Contains load/shipment information including:",
        "• Pickup and delivery locations",
        "• Cargo descriptions and weights",
        "• Driver and equipment assignments",
        "• Scheduling and rate information",
        "",
        "⚙️ SETTINGS SHEET",
        "Contains company configuration including:",
        "• Company information",
        "• System preferences",
        "• Default values and prefixes",
        "• Notification settings",
        "",
        "🔍 DATA VALIDATION",
        "Before importing, ensure:",
        "• All required fields (*) are completed",
        "• Dates are valid and properly formatted",
        "• Phone numbers are 10 digits",
        "• Email addresses are valid",
        "• License plates are unique",
        "• VIN numbers are 17 characters",
        "• Load numbers are unique",
        "",
        "📤 IMPORT PROCESS",
        "1. Complete all relevant sheets with your data",
        "2. Use the Validation sheet to check for errors",
        "3. Save as Excel format (.xlsx)",
        "4. Upload through Launch platform import feature",
        "5. Review import summary",
        "6. Confirm import to add data to system",
        "",
        "❓ SUPPORT",
        "For assistance, contact your system administrator",
        "or refer to the Launch platform documentation.",
        "",
        f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ]
    
    # Add instructions to sheet
    for row_num, instruction in enumerate(instructions, 1):
        ws[f'A{row_num}'] = instruction
        
        # Format headers and important text
        if instruction.startswith('🚀') or instruction.startswith('📋') or instruction.startswith('📝'):
            ws[f'A{row_num}'].font = Font(bold=True, size=14, color='1976D2')
        elif instruction.startswith('📊') or instruction.startswith('🔍') or instruction.startswith('📤'):
            ws[f'A{row_num}'].font = Font(bold=True, size=12, color='1976D2')
        elif instruction.startswith('🧑‍💼') or instruction.startswith('🚛') or instruction.startswith('🚚') or instruction.startswith('📦') or instruction.startswith('⚙️'):
            ws[f'A{row_num}'].font = Font(bold=True, size=11, color='D32F2F')
        elif instruction.startswith('•'):
            ws[f'A{row_num}'].font = Font(size=10)
    
    # Set column width
    ws.column_dimensions['A'].width = 80

def create_validation_sheet(wb):
    """Create a validation sheet with status value references."""
    ws = wb.create_sheet(title="Validation")
    
    # Add title
    ws.merge_cells('A1:D1')
    ws['A1'] = "🔍 Data Validation Reference"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add spacing
    ws.append([])
    ws.append([])
    
    # Status values section
    ws['A3'] = "📊 Valid Status Values"
    ws['A3'].font = Font(bold=True, size=12, color='1976D2')
    
    ws.append([])
    
    # Driver Status
    ws.append(['Driver Status', 'Description', '', ''])
    ws['A5'].font = Font(bold=True)
    ws['B5'].font = Font(bold=True)
    
    driver_statuses = [
        ['active', 'Driver is actively working'],
        ['in-training', 'Driver is in training period'],
        ['oos', 'Driver is out of service']
    ]
    
    for status in driver_statuses:
        ws.append([status[0], status[1], '', ''])
    
    ws.append(['', '', '', ''])
    
    # Truck/Trailer Status
    current_row = ws.max_row + 1
    ws[f'A{current_row}'] = 'Truck/Trailer Status'
    ws[f'B{current_row}'] = 'Description'
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'B{current_row}'].font = Font(bold=True)
    
    truck_statuses = [
        ['available', 'Vehicle is available for assignment'],
        ['in-use', 'Vehicle is currently assigned/in use'],
        ['maintenance', 'Vehicle is undergoing maintenance'],
        ['out-of-service', 'Vehicle is out of service']
    ]
    
    for status in truck_statuses:
        ws.append([status[0], status[1], '', ''])
    
    ws.append(['', '', '', ''])
    
    # Load Status
    current_row = ws.max_row + 1
    ws[f'A{current_row}'] = 'Load Status'
    ws[f'B{current_row}'] = 'Description'
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'B{current_row}'].font = Font(bold=True)
    
    load_statuses = [
        ['pending', 'Load is awaiting assignment'],
        ['assigned', 'Load has been assigned to driver/truck'],
        ['picked-up', 'Load has been picked up'],
        ['in-transit', 'Load is en route to destination'],
        ['delivered', 'Load has been delivered'],
        ['cancelled', 'Load has been cancelled']
    ]
    
    for status in load_statuses:
        ws.append([status[0], status[1], '', ''])
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50

if __name__ == "__main__":
    create_excel_template()
