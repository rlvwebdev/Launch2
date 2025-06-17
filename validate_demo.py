#!/usr/bin/env python3
"""
Validate the demo data file
"""
import openpyxl

def validate_demo_file():
    try:
        wb = openpyxl.load_workbook('Launch_Demo_Import_Data.xlsx')
        
        print("📊 Demo Data File Validation:")
        print(f"   📑 Total Sheets: {len(wb.sheetnames)}")
        print(f"   📋 Sheet Names: {', '.join(wb.sheetnames)}")
        print()
          # Count records in each sheet (subtract 3: description row, empty row, header row)
        drivers_count = wb["Drivers"].max_row - 3
        trucks_count = wb["Trucks"].max_row - 3
        trailers_count = wb["Trailers"].max_row - 3
        loads_count = wb["Loads"].max_row - 3
        settings_count = wb["Settings"].max_row - 3
        
        print("📊 Record Counts:")
        print(f"   👨‍💼 Drivers: {drivers_count} records")
        print(f"   🚛 Trucks: {trucks_count} records")
        print(f"   🚚 Trailers: {trailers_count} records")
        print(f"   📦 Loads: {loads_count} records")
        print(f"   ⚙️ Settings: {settings_count} records")
        print()
        
        # Validate target counts
        target_counts = {
            'Drivers': 23,
            'Trucks': 42,
            'Trailers': 63,
            'Loads': 119,
            'Settings': 20
        }
        
        actual_counts = {
            'Drivers': drivers_count,
            'Trucks': trucks_count,
            'Trailers': trailers_count,
            'Loads': loads_count,
            'Settings': settings_count
        }
        
        print("✅ Target vs Actual:")
        all_correct = True
        for entity, target in target_counts.items():
            actual = actual_counts[entity]
            status = "✅" if actual == target else "❌"
            print(f"   {status} {entity}: {actual}/{target}")
            if actual != target:
                all_correct = False
        
        print()
        if all_correct:
            print("🎉 All record counts match targets!")
        else:
            print("⚠️ Some record counts don't match targets")
            
        # Sample some data
        print("\n📋 Sample Data Preview:")
        
        # Sample driver
        driver_row = wb["Drivers"][4]  # First data row
        print(f"   First Driver: {driver_row[1].value} {driver_row[2].value} ({driver_row[0].value})")
        
        # Sample truck
        truck_row = wb["Trucks"][4]  # First data row
        print(f"   First Truck: {truck_row[5].value} {truck_row[1].value} {truck_row[2].value} ({truck_row[0].value})")
        
        # Sample load
        load_row = wb["Loads"][4]  # First data row
        print(f"   First Load: {load_row[1].value} from {load_row[5].value}, {load_row[6].value} to {load_row[9].value}, {load_row[10].value}")
        
        wb.close()
        print("\n✅ Demo data file validation complete!")
        
    except Exception as e:
        print(f"❌ Error validating demo file: {e}")

if __name__ == "__main__":
    validate_demo_file()
